import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from msme_tool.config import load_config, write_default_config
from msme_tool.process import process_file
from msme_tool.report import save_report

FIXTURES = Path(__file__).parent / "fixtures"
GREEN = str(FIXTURES / "Green Wood.xls")
PUNJAB = str(FIXTURES / "Punjab Plywood.xls")


def test_config_overrides_and_as_on(tmp_path):
    cfg_path = tmp_path / "config.json"
    cfg_path.write_text(json.dumps({
        "default_agreed_days": 45,
        "per_party_agreed_days": {"Green Wood": 30},
        "as_on_date": "2026-03-31",
    }))
    cfg = load_config(cfg_path)
    assert cfg.agreed_days_for("M/s. Green Wood Crafts") == 30
    assert cfg.agreed_days_for("Someone Else") == 45
    assert cfg.as_on_for(date(2025, 1, 1)) == date(2026, 3, 31)


def test_shorter_window_increases_disallowance(tmp_path):
    cfg_path = tmp_path / "config.json"
    cfg_path.write_text(json.dumps({"default_agreed_days": 15}))
    cfg15 = load_config(cfg_path)
    cfg45 = load_config(None)
    dis15 = process_file(PUNJAB, cfg15).disallowance.total_disallowed
    dis45 = process_file(PUNJAB, cfg45).disallowance.total_disallowed
    assert dis15 >= dis45  # a tighter window can only disallow more


def test_write_default_config_roundtrips(tmp_path):
    p = tmp_path / "config.json"
    write_default_config(p)
    cfg = load_config(p)
    assert cfg.default_agreed_days == 45
    assert cfg.rate_schedule().rate_on(date(2026, 1, 1)) > 0


def test_report_writes_expected_sheets(tmp_path):
    cfg = load_config(None)
    results = [process_file(GREEN, cfg), process_file(PUNJAB, cfg)]
    out = tmp_path / "report.xlsx"
    save_report(results, str(out))
    assert out.exists()

    wb = load_workbook(str(out))
    assert "Summary" in wb.sheetnames
    assert "Flagged" in wb.sheetnames
    # a per-party sheet exists for each result
    assert len(wb.sheetnames) >= 4
