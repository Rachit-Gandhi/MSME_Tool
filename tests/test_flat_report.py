from pathlib import Path

import pytest
from openpyxl import load_workbook

from msme_tool.config import load_config
from msme_tool.disallowance import STATUS_DISALLOWED
from msme_tool.process import process_file
from msme_tool.report import _FLAT_HEADERS, build_flat_workbook, save_flat_report

FIXTURES = Path(__file__).parent / "fixtures"
GREEN = str(FIXTURES / "Green Wood.xls")
PUNJAB = str(FIXTURES / "Punjab Plywood.xls")

_HDR_ROW = 3
_FIRST_DATA_ROW = _HDR_ROW + 1


@pytest.fixture
def results():
    cfg = load_config(None)
    return [process_file(GREEN, cfg), process_file(PUNJAB, cfg)]


def test_flat_headers_and_total_row(results, tmp_path):
    out = tmp_path / "flat.xlsx"
    save_flat_report(results, str(out))
    assert out.exists()

    ws = load_workbook(str(out))["Items"]
    headers = [ws.cell(row=_HDR_ROW, column=c).value for c in range(1, len(_FLAT_HEADERS) + 1)]
    assert headers == _FLAT_HEADERS

    n_items = sum(len(r.disallowance.assessments) for r in results)
    total_row = _FIRST_DATA_ROW + n_items
    assert ws.cell(row=total_row, column=1).value == "TOTAL"

    exp_purchase = round(
        sum(a.item.amount for r in results for a in r.disallowance.assessments), 2
    )
    assert ws.cell(row=total_row, column=3).value == pytest.approx(exp_purchase)


def test_flat_yn_flags_match_assessments(results):
    ws = build_flat_workbook(results)["Items"]

    expected = []
    for r in results:
        pe = r.ledger.period_end
        for a in r.disallowance.assessments:
            expected.append(
                (
                    "Y" if a.appointed_day <= pe else "N",
                    "Y" if a.status == STATUS_DISALLOWED else "N",
                )
            )

    for i, (exp_45, exp_dis) in enumerate(expected):
        row = _FIRST_DATA_ROW + i
        assert ws.cell(row=row, column=6).value == exp_45
        assert ws.cell(row=row, column=7).value == exp_dis

    # The two fixtures give meaningful coverage: Punjab disallows, Green does not.
    dis_flags = {e[1] for e in expected}
    assert dis_flags == {"Y", "N"}
