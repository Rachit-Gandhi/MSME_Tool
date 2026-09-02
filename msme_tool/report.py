"""Write the results to an Excel workbook with openpyxl.

Sheets:
* ``Summary``  -- assumption banner, one row per party, grand totals and the
                  Form 3CD Clause 22 disallowance figure.
* one sheet per party -- the FIFO audit trail (every invoice, its settlements,
                  appointed day, unpaid, disallowed, interest) plus that party's
                  flagged entries.
* ``Flagged``  -- every non-payment entry across all parties, for manual review.
"""

from __future__ import annotations

import re
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .disallowance import STATUS_DISALLOWED
from .process import PartyResult

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_BANNER_FILL = PatternFill("solid", fgColor="FFF2CC")
_BANNER_FONT = Font(bold=True, color="7F6000")
_MONEY = "#,##0.00"
_DATE = "dd-mmm-yyyy"

_BANNER = (
    "ASSUMPTIONS: every party file is treated as an in-scope Micro/Small supplier "
    "(Medium/non-MSME suppliers are NOT excluded here - verify separately). "
    "Payment window assumed per config (default 45 days). Payment vouchers, contra "
    "Sales (Tax Invoice) and TDS (Journal to a TDS account) settle invoices under "
    "FIFO as receipts; other Journals/Credit Notes are flagged, not netted. "
    "Opening balances are aged from the period start. "
    "Interest = 3x RBI bank rate, monthly rests - verify the rate schedule."
)


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    clean = re.sub(r"[:\\/?*\[\]]", " ", name).strip()[:28] or "Party"
    candidate = clean
    n = 2
    while candidate.lower() in used:
        candidate = f"{clean[:25]} {n}"
        n += 1
    used.add(candidate.lower())
    return candidate


def _write_header(ws, headers, row=1):
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(results: list[PartyResult]) -> Workbook:
    wb = Workbook()
    _build_summary(wb, results)

    used_names: set[str] = {"summary", "flagged"}
    for r in results:
        _build_party_sheet(wb, r, used_names)

    _build_flagged(wb, results)
    return wb


def _build_summary(wb: Workbook, results: list[PartyResult]) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws.cell(row=1, column=1, value="MSME 43B(h) Disallowance & Section 16 Interest").font = _TITLE_FONT
    banner = ws.cell(row=2, column=1, value=_BANNER)
    banner.fill = _BANNER_FILL
    banner.font = _BANNER_FONT
    banner.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    ws.row_dimensions[2].height = 60

    headers = [
        "Party", "Agreed days", "Period end", "As-on (interest)",
        "Total purchases/opening", "Outstanding at year-end",
        "43B(h) disallowed", "Sec 16 interest", "Recon drift",
    ]
    hdr_row = 4
    _write_header(ws, headers, row=hdr_row)

    tot_open = tot_out = tot_dis = tot_int = 0.0
    r = hdr_row + 1
    for res in results:
        out_at_ye = round(
            sum(a.unpaid_at_period_end for a in res.disallowance.assessments), 2
        )
        dis = res.disallowance.total_disallowed
        inte = res.interest.total_interest
        tot_open += res.fifo.open_total
        tot_out += out_at_ye
        tot_dis += dis
        tot_int += inte

        ws.cell(row=r, column=1, value=res.ledger.party)
        ws.cell(row=r, column=2, value=res.agreed_days)
        ws.cell(row=r, column=3, value=res.ledger.period_end).number_format = _DATE
        ws.cell(row=r, column=4, value=res.as_on).number_format = _DATE
        ws.cell(row=r, column=5, value=res.fifo.open_total).number_format = _MONEY
        ws.cell(row=r, column=6, value=out_at_ye).number_format = _MONEY
        ws.cell(row=r, column=7, value=dis).number_format = _MONEY
        ws.cell(row=r, column=8, value=inte).number_format = _MONEY
        drift = res.reconciliation_drift
        dc = ws.cell(row=r, column=9, value=("n/a" if drift is None else drift))
        if isinstance(drift, (int, float)) and abs(drift) > 0.05:
            dc.font = Font(color="C00000", bold=True)
        r += 1

    # Totals row
    tr = r
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    for col, val in ((5, tot_open), (6, tot_out), (7, tot_dis), (8, tot_int)):
        c = ws.cell(row=tr, column=col, value=round(val, 2))
        c.number_format = _MONEY
        c.font = Font(bold=True)

    # Form 3CD Clause 22 line
    c22 = tr + 2
    ws.cell(row=c22, column=1, value="Form 3CD Clause 22 - total 43B(h) disallowance:").font = Font(bold=True)
    cell = ws.cell(row=c22, column=7, value=round(tot_dis, 2))
    cell.number_format = _MONEY
    cell.font = Font(bold=True, color="C00000")

    _autosize(ws, [42, 11, 14, 16, 20, 20, 18, 16, 12])


def _fmt_settlements(item) -> str:
    if not item.settlements:
        return ""
    return "; ".join(
        f"{s.pay_date:%d-%b-%y}: {s.amount:,.2f} [{s.source}]"
        + (f" (#{s.pay_vch_no})" if s.pay_vch_no else "")
        for s in item.settlements
    )


def _build_party_sheet(wb: Workbook, res: PartyResult, used: set[str]) -> None:
    ws = wb.create_sheet(_sanitize_sheet_name(res.ledger.party, used))

    ws.cell(row=1, column=1, value=res.ledger.party).font = _TITLE_FONT
    meta = (
        f"Period {res.ledger.period_start:%d-%b-%Y} to {res.ledger.period_end:%d-%b-%Y}  |  "
        f"Agreed days: {res.agreed_days}  |  Interest as-on: {res.as_on:%d-%b-%Y}  |  "
        f"File closing: {res.ledger.closing_balance}  |  Recon drift: {res.reconciliation_drift}"
    )
    ws.cell(row=2, column=1, value=meta)

    fifo = res.fifo
    settle_meta = (
        f"Receipts settling invoices (FIFO)  |  Payments: {fifo.payment_total:,.2f}  |  "
        f"TDS: {fifo.tds_total:,.2f}  |  Sales: {fifo.sales_total:,.2f}  |  "
        f"Total: {fifo.settlement_total:,.2f}"
    )
    ws.cell(row=3, column=1, value=settle_meta).font = Font(italic=True, color="1F4E78")

    headers = [
        "Inv date", "Vch no", "Type", "Amount", "Appointed day",
        "Settlements (FIFO)", "Paid by YE", "Unpaid at YE",
        "Status", "43B(h) disallowed", "Sec 16 interest",
    ]
    hdr = 4
    _write_header(ws, headers, row=hdr)

    interest_by_row = {ii.item.row: ii for ii in res.interest.items}
    r = hdr + 1
    for a in res.disallowance.assessments:
        item = a.item
        ii = interest_by_row.get(item.row)
        ws.cell(row=r, column=1, value=item.date).number_format = _DATE
        ws.cell(row=r, column=2, value=item.vch_no or ("OPENING" if item.is_opening_balance else ""))
        ws.cell(row=r, column=3, value="Opening Balance" if item.is_opening_balance else "Purchase")
        ws.cell(row=r, column=4, value=item.amount).number_format = _MONEY
        ws.cell(row=r, column=5, value=a.appointed_day).number_format = _DATE
        ws.cell(row=r, column=6, value=_fmt_settlements(item))
        ws.cell(row=r, column=7, value=round(item.paid_upto(res.ledger.period_end), 2)).number_format = _MONEY
        ws.cell(row=r, column=8, value=a.unpaid_at_period_end).number_format = _MONEY
        sc = ws.cell(row=r, column=9, value=a.status)
        if a.status == STATUS_DISALLOWED:
            sc.font = Font(color="C00000", bold=True)
        ws.cell(row=r, column=10, value=a.disallowed).number_format = _MONEY
        ws.cell(row=r, column=11, value=(ii.total if ii else 0.0)).number_format = _MONEY
        r += 1

    # party totals
    ws.cell(row=r, column=3, value="TOTAL").font = Font(bold=True)
    for col, val in (
        (10, res.disallowance.total_disallowed),
        (11, res.interest.total_interest),
    ):
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = _MONEY
        c.font = Font(bold=True)

    # flagged entries for this party
    if res.fifo.flagged:
        fr = r + 3
        ws.cell(row=fr, column=1, value="Flagged entries (not auto-netted - review manually)").font = Font(bold=True, color="7F6000")
        _write_header(ws, ["Date", "Type", "Vch no", "Account", "Debit", "Credit", "Reason"], row=fr + 1)
        rr = fr + 2
        for f in res.fifo.flagged:
            ws.cell(row=rr, column=1, value=f.date).number_format = _DATE
            ws.cell(row=rr, column=2, value=f.vch_type)
            ws.cell(row=rr, column=3, value=f.vch_no)
            ws.cell(row=rr, column=4, value=f.account)
            ws.cell(row=rr, column=5, value=f.debit).number_format = _MONEY
            ws.cell(row=rr, column=6, value=f.credit).number_format = _MONEY
            ws.cell(row=rr, column=7, value=f.reason)
            rr += 1

    _autosize(ws, [12, 18, 15, 16, 14, 34, 14, 14, 14, 18, 16])


def _build_flagged(wb: Workbook, results: list[PartyResult]) -> None:
    ws = wb.create_sheet("Flagged")
    ws.cell(row=1, column=1, value="Flagged entries across all parties (manual review)").font = _TITLE_FONT
    _write_header(ws, ["Party", "Date", "Type", "Vch no", "Account", "Debit", "Credit", "Reason"], row=3)
    r = 4
    for res in results:
        for f in res.fifo.flagged:
            ws.cell(row=r, column=1, value=res.ledger.party)
            ws.cell(row=r, column=2, value=f.date).number_format = _DATE
            ws.cell(row=r, column=3, value=f.vch_type)
            ws.cell(row=r, column=4, value=f.vch_no)
            ws.cell(row=r, column=5, value=f.account)
            ws.cell(row=r, column=6, value=f.debit).number_format = _MONEY
            ws.cell(row=r, column=7, value=f.credit).number_format = _MONEY
            ws.cell(row=r, column=8, value=f.reason)
            r += 1
    _autosize(ws, [42, 12, 15, 15, 30, 14, 14, 40])


def save_report(results: list[PartyResult], out_path: str) -> str:
    wb = build_workbook(results)
    wb.save(out_path)
    return out_path


# --- flat one-row-per-invoice export -------------------------------------------

_FLAT_HEADERS = [
    "Party", "Date of Purchase", "Purchase Amt", "Date of Payments",
    "Interest Accrued", "45 Days (Y/N)", "Disallowed (Y/N)",
    "Paid Amt", "Pending Amt",
]


def _fmt_payment_dates(item) -> str:
    """Semicolon-joined settlement dates for the flat export (blank if none)."""
    return "; ".join(f"{s.pay_date:%d-%b-%Y}" for s in item.settlements)


def build_flat_workbook(results: list[PartyResult]) -> Workbook:
    """One sheet, one row per open item (invoice / opening balance) across parties.

    ``45 Days (Y/N)`` is age/window based: ``Y`` when the item's payment window has
    elapsed by year-end (appointed day <= period end), regardless of payment timing.
    ``Disallowed (Y/N)`` reflects the §43B(h) status (still unpaid past the window
    at year-end).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    ws.cell(row=1, column=1, value="MSME 43B(h) - invoice-level summary").font = _TITLE_FONT
    hdr = 3
    _write_header(ws, _FLAT_HEADERS, row=hdr)

    tot_purchase = tot_interest = tot_paid = tot_pending = 0.0
    r = hdr + 1
    for res in results:
        period_end = res.ledger.period_end
        interest_by_row = {ii.item.row: ii for ii in res.interest.items}
        for a in res.disallowance.assessments:
            item = a.item
            ii = interest_by_row.get(item.row)
            interest = ii.total if ii else 0.0
            paid = round(item.paid_upto(period_end), 2)
            pending = a.unpaid_at_period_end

            ws.cell(row=r, column=1, value=res.ledger.party)
            ws.cell(row=r, column=2, value=item.date).number_format = _DATE
            ws.cell(row=r, column=3, value=item.amount).number_format = _MONEY
            ws.cell(row=r, column=4, value=_fmt_payment_dates(item))
            ws.cell(row=r, column=5, value=interest).number_format = _MONEY
            ws.cell(row=r, column=6, value=("Y" if a.appointed_day <= period_end else "N"))
            dc = ws.cell(row=r, column=7, value=("Y" if a.status == STATUS_DISALLOWED else "N"))
            if a.status == STATUS_DISALLOWED:
                dc.font = Font(color="C00000", bold=True)
            ws.cell(row=r, column=8, value=paid).number_format = _MONEY
            ws.cell(row=r, column=9, value=pending).number_format = _MONEY

            tot_purchase += item.amount
            tot_interest += interest
            tot_paid += paid
            tot_pending += pending
            r += 1

    # TOTAL row
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for col, val in ((3, tot_purchase), (5, tot_interest), (8, tot_paid), (9, tot_pending)):
        c = ws.cell(row=r, column=col, value=round(val, 2))
        c.number_format = _MONEY
        c.font = Font(bold=True)

    _autosize(ws, [42, 16, 16, 34, 16, 13, 15, 16, 16])
    return wb


def save_flat_report(results: list[PartyResult], out_path: str) -> str:
    wb = build_flat_workbook(results)
    wb.save(out_path)
    return out_path
