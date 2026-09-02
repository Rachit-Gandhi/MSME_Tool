"""Read a Tally ledger export into a :class:`~msme_tool.model.Ledger`.

Key robustness points (all confirmed against real exports):

* The files are OOXML/zip (``.xlsx`` content) even when named ``.xls``. openpyxl
  refuses the ``.xls`` *extension*, so we detect the ``PK`` zip magic and feed the
  bytes through :class:`io.BytesIO`, which bypasses the extension check.
* The layout is a Tally ledger: a letterhead block, the ledger/party name, a
  period line (``1-Apr-25 to 31-Mar-26``), a header row where column A == "Date",
  transaction rows, then total / closing-balance rows that must be skipped.
* Dates come back as ``datetime``; amounts as ``int``/``float`` (occasionally an
  empty string). Both are coerced defensively.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime

import openpyxl

from .model import Ledger, Transaction

# Tally often labels the opening balance row with no Vch Type; we synthesise one.
_OPENING_BALANCE_ACCOUNT = "opening balance"
_CLOSING_BALANCE_ACCOUNT = "closing balance"

_PERIOD_RE = re.compile(
    r"(\d{1,2}-[A-Za-z]{3}-\d{2,4})\s*to\s*(\d{1,2}-[A-Za-z]{3}-\d{2,4})"
)


class LedgerParseError(ValueError):
    """Raised when a file cannot be interpreted as a Tally ledger export."""


def _to_float(value) -> float:
    """Coerce a cell to float; blanks/None -> 0.0; strip commas and currency."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if s in ("", "-"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _to_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Excel serial fallback (1900 date system) for the rare text/serial case.
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + _timedelta_days(value)).date()
        except (OverflowError, ValueError):
            return None
    return None


def _timedelta_days(n):
    from datetime import timedelta

    return timedelta(days=float(n))


def _parse_period(text: str) -> tuple[date, date]:
    m = _PERIOD_RE.search(text)
    if not m:
        raise LedgerParseError(f"could not parse period from {text!r}")
    start = _parse_dmy(m.group(1))
    end = _parse_dmy(m.group(2))
    return start, end


def _parse_dmy(token: str) -> date:
    for fmt in ("%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    raise LedgerParseError(f"could not parse date token {token!r}")


def load_workbook_any(path: str):
    """Open ``path`` as an xlsx workbook regardless of its file extension."""
    with open(path, "rb") as fh:
        data = fh.read()
    if data[:2] != b"PK":
        raise LedgerParseError(
            f"{path!r} is not OOXML/xlsx content (true legacy .xls is unsupported; "
            "re-export from Tally as Excel/xlsx)"
        )
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)


def read_ledger(path: str) -> Ledger:
    """Parse a single Tally ledger export file into a :class:`Ledger`."""
    wb = load_workbook_any(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    party = _extract_party(rows)
    period_start, period_end = _extract_period(rows)
    header_idx = _find_header(rows)

    ledger = Ledger(
        party=party,
        period_start=period_start,
        period_end=period_end,
        source_file=path,
    )

    for offset, raw in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        txn = _parse_row(raw, offset, period_start)
        if txn is None:
            # Could be a closing-balance / totals row: capture closing balance.
            _maybe_capture_closing(raw, ledger)
            continue
        ledger.transactions.append(txn)

    return ledger


def _extract_party(rows) -> str:
    """The ledger/party name sits just above the 'Ledger Account' label row."""
    for i, r in enumerate(rows[:15]):
        first = _cell(r, 0)
        if isinstance(first, str) and first.strip().lower() == "ledger account":
            # party name is the non-empty line immediately above the label
            for j in range(i - 1, -1, -1):
                cand = _cell(rows[j], 0)
                if isinstance(cand, str) and cand.strip():
                    return cand.strip()
    # Fallback: first non-empty header line.
    for r in rows[:15]:
        cand = _cell(r, 0)
        if isinstance(cand, str) and cand.strip():
            return cand.strip()
    return "UNKNOWN PARTY"


def _extract_period(rows) -> tuple[date, date]:
    for r in rows[:15]:
        cand = _cell(r, 0)
        if isinstance(cand, str) and _PERIOD_RE.search(cand):
            return _parse_period(cand)
    raise LedgerParseError("could not locate the period line (e.g. '1-Apr-25 to 31-Mar-26')")


def _find_header(rows) -> int:
    for i, r in enumerate(rows):
        if isinstance(_cell(r, 0), str) and _cell(r, 0).strip().lower() == "date":
            return i
    raise LedgerParseError("could not locate the transaction header row (col A == 'Date')")


def _parse_row(raw, row_no: int, period_start: date) -> Transaction | None:
    """Turn a data row into a Transaction, or None if it is a totals/blank row."""
    dt = _to_date(_cell(raw, 0))
    particulars = _cell(raw, 1)
    account = _cell(raw, 2)
    vch_type = _cell(raw, 3)
    vch_no = _cell(raw, 4)
    debit = _to_float(_cell(raw, 5))
    credit = _to_float(_cell(raw, 6))

    account_l = (account or "").strip().lower() if isinstance(account, str) else ""

    # Skip the closing-balance line and pure totals rows (no By/To marker).
    if account_l == _CLOSING_BALANCE_ACCOUNT:
        return None
    particulars_s = particulars.strip() if isinstance(particulars, str) else ""
    if particulars_s not in ("By", "To"):
        # Totals / grand-total rows have amounts but no By/To marker.
        return None
    if dt is None:
        return None

    # Opening balance rows carry no Vch Type in Tally -> synthesise one so the
    # classifier treats them as open items dated at period start.
    if account_l == _OPENING_BALANCE_ACCOUNT and not (vch_type or "").strip():
        vch_type = "Opening Balance"
        # Real invoice dates predate the year; age from period start (agreed rule).
        dt = period_start

    return Transaction(
        date=dt,
        particulars=particulars_s,
        account=account.strip() if isinstance(account, str) else account,
        vch_type=vch_type.strip() if isinstance(vch_type, str) else vch_type,
        vch_no=str(vch_no).strip() if vch_no not in (None, "") else None,
        debit=debit,
        credit=credit,
        row=row_no,
    )


def _maybe_capture_closing(raw, ledger: Ledger) -> None:
    account = _cell(raw, 2)
    if isinstance(account, str) and account.strip().lower() == _CLOSING_BALANCE_ACCOUNT:
        # Tally prints the closing balance in the opposite column; capture both.
        val = _to_float(_cell(raw, 5)) or _to_float(_cell(raw, 6)) or _to_float(_cell(raw, 3))
        if val:
            ledger.closing_balance = val


def _cell(row_tuple, idx: int):
    if row_tuple is None or idx >= len(row_tuple):
        return None
    return row_tuple[idx]
